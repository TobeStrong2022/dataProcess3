# -*- coding: utf-8 -*-
# @Author   : zilong
# @Time     : 2024/12/24 10:38
# 一是通过对卷烟社会存销比、动销率、库存周转率等5个指标的监控，构建市场预警模块和响应机制，掌握市场消费特征消费趋势变化和调整营销决策。
# 存销比 = 月末库存金额 / 当月销售额
# 动销售率 = 本月已销品牌数 / 店铺经营的商品总品种数
# 库存周转率 = 该期间的出库总金额 / 该期间的平均库存金额
# 条包比例 = 条笔数 / 包笔数
# 单笔销售额 = 总销售额 / 总销售笔数

# 二是围绕客户的基础信息、主销规格、品牌宽度等5类数据，全面掌握、准确诊断客户经营情况，提供针对性经营指导。

import pandas as pd
import shutil
import time
import openpyxl
from openpyxl.styles import Font  # 导入字体模块
from openpyxl.utils.dataframe import dataframe_to_rows

# 任务一：掌握市场消费特征消费趋势变化和调整营销决策
# 1.读取数据并完成初步数据清洗
print("处理任务一...")
print("读取数据并完成初步数据清洗...")
data_previous = pd.read_excel('原始数据/基准数据（上月）.xlsx')
# 删除含有“-”的行
data_previous = data_previous[data_previous['品牌数'] != '-']
data_previous = data_previous[data_previous['条包比例'] != '-']
data_previous = data_previous[data_previous['笔均销量'] != '-']
data_previous = data_previous[data_previous['库存数量'] != '-']
data_previous = data_previous[data_previous['初期库存'] != '-']
data_previous = data_previous[data_previous['库存金额'] != '-']
data_previous = data_previous[data_previous['卷烟销售额'] != '-']

data_current = pd.read_excel('原始数据/当前数据（本月）.xlsx')
# 删除含有“-”的行
data_current = data_current[data_current['品牌数'] != '-']
data_current = data_current[data_current['条包比例'] != '-']
data_current = data_current[data_current['笔均销量'] != '-']
data_current = data_current[data_current['库存数量'] != '-']
data_current = data_current[data_current['初期库存'] != '-']
data_current = data_current[data_current['库存金额'] != '-']
data_current = data_current[data_current['卷烟销售额'] != '-']

# 2.数据计算
print("数据计算...")
# 2.1计算存销比
print("计算存销比...")
data_previous['存销比'] = data_previous['库存金额'] / data_previous['卷烟销售额']
data_current['存销比'] = data_current['库存金额'] / data_current['卷烟销售额']
# 2.2计算动销率
print("计算动销率...")
data_previous['动销率'] = data_previous['本月已销品牌数'] / data_previous['品牌数']
data_current['动销率'] = data_current['本月已销品牌数'] / data_current['品牌数']
# 2.3计算库存周转率
print("计算库存周转率...")
data_previous['库存周转率'] = data_previous['总销量'] / ((data_previous['初期库存'] + data_previous['库存数量']) / 2)
data_current['库存周转率'] = data_current['总销量'] / ((data_current['初期库存'] + data_current['库存数量']) / 2)
# 2.4计算条包比例
print("计算条包比例...")
# 原数据已有，无需计算
# 2.5计算单笔销售额
print("计算单笔销售额...")
data_previous['单笔销售额'] = data_previous['卷烟销售额'] / data_previous['总销量']
data_current['单笔销售额'] = data_current['卷烟销售额'] / data_current['总销量']

# 3.数据分析
# 根据五项指标的变化情况进行预警分析，分别计算两个表格中五项指标的平均值
# 若data_current对比data_previous五项指标中，平均值波动超过预警比例则进行预警提示
warningsRatio = 0.25
print("数据分析...")
# 计算存销比、动销率、库存周转率、条包比例、单笔销售额的平均值
# 存销比均值
average_previous_storage_sales_ratio = data_previous['存销比'].mean()
average_current_storage_sales_ratio = data_current['存销比'].mean()
# 动销率均值
average_previous_turnover_rate = data_previous['动销率'].mean()
average_current_turnover_rate = data_current['动销率'].mean()
# 库存周转率均值
average_previous_inventory_turnover_rate = data_previous['库存周转率'].mean()
average_current_inventory_turnover_rate = data_current['库存周转率'].mean()
# 条包比例均值
average_previous_tiaobao_ratio = data_previous['条包比例'].mean()
average_current_tiaobao_ratio = data_current['条包比例'].mean()
# 单笔销售额均值
average_previous_single_sales = data_previous['单笔销售额'].mean()
average_current_single_sales = data_current['单笔销售额'].mean()
# 计算平均值波动
storage_sales_ratio_fluctuation = (average_current_storage_sales_ratio - average_previous_storage_sales_ratio) / average_previous_storage_sales_ratio
turnover_rate_fluctuation = (average_current_turnover_rate - average_previous_turnover_rate) / average_previous_turnover_rate
inventory_turnover_rate_fluctuation = (average_current_inventory_turnover_rate - average_previous_inventory_turnover_rate) / average_previous_inventory_turnover_rate
tiaobao_ratio_fluctuation = (average_current_tiaobao_ratio - average_previous_tiaobao_ratio) / average_previous_tiaobao_ratio
single_sales_fluctuation = (average_current_single_sales - average_previous_single_sales) / average_previous_single_sales

# 4.预警提示
print("预警提示文件生成...")
# 复制模版文件
# 生成时间戳
time_stamp = time.strftime('%d%H%M%S', time.localtime(time.time()))
shutil.copyfile('原始数据/预警提示模版.xlsx', '分析结果/预警提示'+time_stamp+'.xlsx')
# 用openpyxl打开文件并写入数据
wb = openpyxl.load_workbook('分析结果/预警提示'+time_stamp+'.xlsx')
ws = wb.active
Color = ['ffc7ce', '9c0006']
font = Font('宋体', size=11, bold=True, italic=False, strike=False, color=Color[1])  # 设置字体样式
suggestStr = ''
ws['B2'] = round(average_previous_storage_sales_ratio,2)
ws['B3'] = round(average_current_storage_sales_ratio,2)
ws['B4'] = str(round(float(storage_sales_ratio_fluctuation) * 100,2)) + '%'
ws['B4'].alignment = openpyxl.styles.Alignment(horizontal='right')  # 设置单元格右对齐
if storage_sales_ratio_fluctuation > warningsRatio or storage_sales_ratio_fluctuation < -warningsRatio:
    suggestStr += '存销比波动过大!'
    # 设置字体颜色为红色
    ws['B4'].font = font
ws['C2'] = round(average_previous_turnover_rate,2)
ws['C3'] = round(average_current_turnover_rate,2)
ws['C4'] = str(round(turnover_rate_fluctuation * 100,2)) + '%'
ws['C4'].alignment = openpyxl.styles.Alignment(horizontal='right')  # 设置单元格右对齐
if turnover_rate_fluctuation > warningsRatio or turnover_rate_fluctuation < -warningsRatio:
    suggestStr += '动销率波动过大!'
    # 设置字体颜色为红色
    ws['C4'].font = font

ws['D2'] = round(average_previous_inventory_turnover_rate,2)
ws['D3'] = round(average_current_inventory_turnover_rate,2)
ws['D4'] = str(round(inventory_turnover_rate_fluctuation * 100,2)) + '%'
ws['D4'].alignment = openpyxl.styles.Alignment(horizontal='right')  # 设置单元格右对齐
if inventory_turnover_rate_fluctuation > warningsRatio or inventory_turnover_rate_fluctuation < -warningsRatio:
    suggestStr += '库存周转率波动过大!'
    # 设置字体颜色为红色
    ws['D4'].font = font

ws['E2'] = round(average_previous_tiaobao_ratio,2)
ws['E3'] = round(average_current_tiaobao_ratio,2)
ws['E4'] = str(round(tiaobao_ratio_fluctuation * 100,2)) + '%'
ws['E4'].alignment = openpyxl.styles.Alignment(horizontal='right')  # 设置单元格右对齐
if tiaobao_ratio_fluctuation > warningsRatio or tiaobao_ratio_fluctuation < -warningsRatio:
    suggestStr += '条包比例波动过大!'
    # 设置字体颜色为红色
    ws['E4'].font = font
ws['F2'] = round(average_previous_single_sales,2)
ws['F3'] = round(average_current_single_sales,2)
ws['F4'] = str(round(single_sales_fluctuation * 100,2)) + '%'
ws['F4'].alignment = openpyxl.styles.Alignment(horizontal='right')  # 设置单元格右对齐
if single_sales_fluctuation > warningsRatio or single_sales_fluctuation < -warningsRatio:
    suggestStr += '单笔销售额波动过大!'
    # 设置字体颜色为红色
    ws['F4'].font = font
ws['A5'] = suggestStr

# 将运算的过程数据data_current和data_previous写入本文件的Sheet2和Sheet3
ws2 = wb.create_sheet('基准数据（上月）')
ws3 = wb.create_sheet('当前数据（本月）')
# 保留前六列和存销比、动销率、库存周转率、条包比例、单笔销售额列
data_previous = pd.concat([data_previous.iloc[:, :6],  data_previous[['存销比', '动销率', '库存周转率', '条包比例', '单笔销售额']],], axis=1)
data_current = pd.concat([data_current.iloc[:, :6],  data_current[['存销比', '动销率', '库存周转率', '条包比例', '单笔销售额']],], axis=1)

for r in dataframe_to_rows(data_previous, index=False, header=True):
    ws2.append(r)
for r in dataframe_to_rows(data_current, index=False, header=True):
    ws3.append(r)

wb.save('分析结果/预警提示'+time_stamp+'.xlsx')


















